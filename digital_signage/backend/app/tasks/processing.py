"""
Zadania przetwarzania treĹ›ci (Celery tasks)
"""
import os
from pathlib import Path
from sqlalchemy.orm import Session

from app.celery_app import celery_app
from app.database import SessionLocal
from app.models.content import Content
from app.models.processing_job import ProcessingJob
from app.config import settings
from app.services.content_service import update_processing_job


@celery_app.task(name="process_image")
def process_image_task(content_id: int, job_id: int):
    """Przetwarzanie obrazu - generowanie miniatury"""
    db: Session = SessionLocal()
    try:
        content = db.query(Content).filter(Content.id == content_id).first()
        if not content:
            update_processing_job(db, job_id, "failed", error_message="Content not found")
            return
        
        update_processing_job(db, job_id, "processing", progress=10)
        
        from PIL import Image
        
        # Otwarcie obrazu
        img = Image.open(content.file_path)
        update_processing_job(db, job_id, "processing", progress=30)
        
        # Generowanie miniatury (max 300x300)
        img.thumbnail((300, 300), Image.Resampling.LANCZOS)
        update_processing_job(db, job_id, "processing", progress=60)
        
        # Zapis miniatury
        os.makedirs(settings.CONTENT_THUMBNAILS_DIR, exist_ok=True)
        thumbnail_path = os.path.join(
            settings.CONTENT_THUMBNAILS_DIR,
            f"thumb_{content_id}.jpg"
        )
        img.convert("RGB").save(thumbnail_path, "JPEG", quality=85)
        update_processing_job(db, job_id, "processing", progress=80)
        
        # Aktualizacja treĹ›ci
        content.thumbnail_path = thumbnail_path
        db.commit()
        
        update_processing_job(db, job_id, "completed", progress=100)
        
    except Exception as e:
        update_processing_job(db, job_id, "failed", error_message=str(e))
    finally:
        db.close()


@celery_app.task(name="process_pdf")
def process_pdf_task(content_id: int, job_id: int):
    """Przetwarzanie PDF - generowanie miniatury z pierwszej strony"""
    db: Session = SessionLocal()
    try:
        content = db.query(Content).filter(Content.id == content_id).first()
        if not content:
            update_processing_job(db, job_id, "failed", error_message="Content not found")
            return
        
        update_processing_job(db, job_id, "processing", progress=10)
        
        from pdf2image import convert_from_path
        from PIL import Image
        
        # Konwersja pierwszej strony do obrazu
        images = convert_from_path(content.file_path, first_page=1, last_page=1)
        update_processing_job(db, job_id, "processing", progress=50)
        
        if images:
            # Generowanie miniatury
            img = images[0]
            img.thumbnail((300, 300), Image.Resampling.LANCZOS)
            update_processing_job(db, job_id, "processing", progress=70)
            
            # Zapis miniatury
            os.makedirs(settings.CONTENT_THUMBNAILS_DIR, exist_ok=True)
            thumbnail_path = os.path.join(
                settings.CONTENT_THUMBNAILS_DIR,
                f"thumb_{content_id}.jpg"
            )
            img.convert("RGB").save(thumbnail_path, "JPEG", quality=85)
            update_processing_job(db, job_id, "processing", progress=90)
            
            # Aktualizacja treĹ›ci
            content.thumbnail_path = thumbnail_path
            db.commit()
        
        update_processing_job(db, job_id, "completed", progress=100)
        
    except Exception as e:
        update_processing_job(db, job_id, "failed", error_message=str(e))
    finally:
        db.close()


@celery_app.task(name="process_excel")
def process_excel_task(content_id: int, job_id: int):
    """Przetwarzanie Excel - analiza struktury"""
    db: Session = SessionLocal()
    try:
        content = db.query(Content).filter(Content.id == content_id).first()
        if not content:
            update_processing_job(db, job_id, "failed", error_message="Content not found")
            return
        
        update_processing_job(db, job_id, "processing", progress=10)
        
        import openpyxl
        
        # Otwarcie pliku Excel
        workbook = openpyxl.load_workbook(content.file_path)
        update_processing_job(db, job_id, "processing", progress=30)
        
        # Analiza struktury
        metadata = {
            "sheets": [],
            "total_rows": 0,
            "total_columns": 0
        }
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            metadata["sheets"].append({
                "name": sheet_name,
                "rows": max_row,
                "columns": max_col
            })
            metadata["total_rows"] += max_row
            metadata["total_columns"] = max(max_col, metadata["total_columns"])
        
        update_processing_job(db, job_id, "processing", progress=80)
        
        # Aktualizacja treĹ›ci
        content.content_metadata = metadata
        db.commit()
        
        update_processing_job(db, job_id, "completed", progress=100)
        
    except Exception as e:
        update_processing_job(db, job_id, "failed", error_message=str(e))
    finally:
        db.close()


@celery_app.task(name="process_video")
def process_video_task(content_id: int, job_id: int):
    """Przetwarzanie video - transkodowanie do MP4 (H.264)"""
    db: Session = SessionLocal()
    try:
        content = db.query(Content).filter(Content.id == content_id).first()
        if not content:
            update_processing_job(db, job_id, "failed", error_message="Content not found")
            return
        
        update_processing_job(db, job_id, "processing", progress=5)
        
        from app.services.video_service import transcode_video, get_video_duration
        
        # ĹšcieĹĽka do przetworzonego pliku
        input_path = content.file_path
        output_filename = f"{content_id}_processed.mp4"
        output_path = os.path.join(settings.CONTENT_PROCESSED_DIR, output_filename)
        
        # Upewnienie siÄ™, ĹĽe katalog istnieje
        os.makedirs(settings.CONTENT_PROCESSED_DIR, exist_ok=True)
        
        update_processing_job(db, job_id, "processing", progress=10)
        
        # Transkodowanie video
        success, error = transcode_video(
            input_path=input_path,
            output_path=output_path,
            width=1920,
            height=1080,
            fps=30,
            crf=23
        )
        
        if not success:
            update_processing_job(db, job_id, "failed", error_message=error)
            return
        
        update_processing_job(db, job_id, "processing", progress=80)
        
        # Pobranie dĹ‚ugoĹ›ci video
        duration = get_video_duration(output_path)
        
        update_processing_job(db, job_id, "processing", progress=90)
        
        # Aktualizacja treĹ›ci
        content.file_path = output_path  # ZastÄ…pienie oryginalnego pliku przetworzonym
        content.video_processed = True
        content.video_format = "mp4"
        if duration:
            content.duration_seconds = duration
        
        db.commit()
        
        update_processing_job(db, job_id, "completed", progress=100)
        
    except Exception as e:
        update_processing_job(db, job_id, "failed", error_message=str(e))
    finally:
        db.close()


