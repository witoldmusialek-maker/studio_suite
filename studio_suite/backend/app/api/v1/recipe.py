"""Service recipe API and generated inventory issues."""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.api.deps import get_current_staff_member, get_current_user, require_salon_access
from app.database import get_db
from app.models.salon_core import (
    Appointment,
    AppointmentPerformedLine,
    InventoryIssue,
    InventoryIssueLine,
    LegacyProductCatalogItem,
    ServiceCatalogItem,
    ServiceRecipeItem,
    StaffMember,
    StockLocation,
)
from app.models.user import User, UserRole
from app.schemas.inventory import (
    ServiceRecipeItemCreate,
    ServiceRecipeItemRead,
    ServiceRecipeItemUpdate,
)

router = APIRouter(tags=["recipe"])


class GeneratedIssueResponse(BaseModel):
    issue_ids: list[int]


class RecipeImportResponse(BaseModel):
    imported_services: int
    imported_rows: int
    skipped_rows: int
    unresolved_products: int


QUANTITY_MODE_LABELS = {
    "EXACT": "Scisle|dokladnie (EXACT)",
    "RANGE": "Zakres (RANGE)",
    "ESTIMATE": "Szacunkowo (ESTIMATE)",
}

INVENTORY_MODE_LABELS = {
    "PER_SERVICE": "Na kazdej usludze (PER_SERVICE)",
    "BATCH_ESTIMATE": "Szacunek zbiorczy (BATCH_ESTIMATE)",
    "STOCKTAKE_ONLY": "Tylko remanent (STOCKTAKE_ONLY)",
}

SERVICE_SEGMENT_LABELS = {
    "PANI": "Pani (PANI)",
    "PAN": "Pan (PAN)",
    "ESTETYKA": "Estetyka (ESTETYKA)",
    "SPRZEDAZ": "Sprzedaz (SPRZEDAZ)",
}

UNIT_ALLOWED_VALUES = ["PCS", "G", "ML", "DOZA"]
BOOL_ALLOWED_VALUES = ["TRUE", "FALSE"]


def _ensure_manager_or_admin(current_user: User) -> None:
    if current_user.role not in {UserRole.ADMIN, UserRole.MANAGER, UserRole.MANAGER_MAIN, UserRole.MANAGER_SALON}:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Only admin/manager can modify recipe")


def _ensure_issue_generation_access(
    db: Session,
    current_user: User,
    current_staff: StaffMember | None,
    appointment: Appointment,
) -> None:
    require_salon_access(db, current_user, appointment.salon_id)
    if current_user.role in {UserRole.ADMIN, UserRole.MANAGER, UserRole.MANAGER_MAIN, UserRole.MANAGER_SALON, UserRole.RECEPTIONIST}:
        return
    if current_user.role != UserRole.EMPLOYEE:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions")
    if current_staff is None:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No staff profile linked to user")


def _estimate_unit_cost(product: LegacyProductCatalogItem | None) -> Decimal:
    if product is None:
        return Decimal("0")
    for candidate in (product.purchase_price, product.purchase_price_c, product.catalog_net_price):
        if candidate is not None:
            return Decimal(str(candidate))
    return Decimal("0")


def _format_decimal(value: Decimal | float | int | None) -> str | None:
    if value is None:
        return None
    number = float(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def _coalesce_decimal(*values: Decimal | None) -> Decimal | None:
    for value in values:
        if value is not None:
            return value
    return None


def _normalize_quantity_mode(value: str | None) -> str:
    normalized = (value or "EXACT").strip().upper()
    if not normalized:
        normalized = "EXACT"
    if "EXACT" in normalized or "SCISLE" in normalized or "DOKLAD" in normalized:
        return "EXACT"
    if "RANGE" in normalized or "ZAKRES" in normalized:
        return "RANGE"
    if "ESTIMATE" in normalized or "SZAC" in normalized:
        return "ESTIMATE"
    if normalized not in {"EXACT", "RANGE", "ESTIMATE"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid quantity_mode")
    return normalized


def _normalize_inventory_mode(value: str | None) -> str:
    normalized = (value or "PER_SERVICE").strip().upper()
    if not normalized:
        normalized = "PER_SERVICE"
    if "PER_SERVICE" in normalized or "NA KAZDEJ USLUDZE" in normalized:
        return "PER_SERVICE"
    if "BATCH_ESTIMATE" in normalized or "SZACUNEK ZBIORCZY" in normalized:
        return "BATCH_ESTIMATE"
    if "STOCKTAKE_ONLY" in normalized or "TYLKO REMANENT" in normalized:
        return "STOCKTAKE_ONLY"
    if normalized not in {"PER_SERVICE", "BATCH_ESTIMATE", "STOCKTAKE_ONLY"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid inventory_mode")
    return normalized


def _normalize_service_segment(value: str | None) -> str:
    normalized = (value or "").strip().upper()
    if not normalized:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid service_segment")
    if "PANI" in normalized:
        return "PANI"
    if normalized.startswith("PAN") or " PAN " in f" {normalized} ":
        return "PAN"
    if "ESTETYKA" in normalized:
        return "ESTETYKA"
    if "SPRZEDAZ" in normalized or "SPRZEDAŻ" in normalized:
        return "SPRZEDAZ"
    raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid service_segment")


def _infer_service_segment_from_code(service_code: str | None) -> str | None:
    code = (service_code or "").strip()
    if not code.isdigit():
        return None
    numeric = int(code)
    if 1 <= numeric <= 99:
        return "PANI"
    if 101 <= numeric <= 199:
        return "PAN"
    if 200 <= numeric <= 299:
        return "SPRZEDAZ"
    if 300 <= numeric <= 399:
        return "ESTETYKA"
    return None


def _resolve_service_segment(service: ServiceCatalogItem) -> str:
    normalized = (service.service_type_code or "").strip().upper()
    if normalized in SERVICE_SEGMENT_LABELS:
        return normalized
    inferred = _infer_service_segment_from_code(service.legacy_code)
    return inferred or "PANI"


def _coerce_decimal(value: float | int | Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return Decimal(str(value))


def _validate_recipe_payload(
    product_family: str | None,
    quantity_mode: str,
    planned_quantity: Decimal | None,
    planned_min: Decimal | None,
    planned_default: Decimal | None,
    planned_max: Decimal | None,
) -> None:
    if not (product_family or "").strip():
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="product_family is required")
    default_value = _coalesce_decimal(planned_default, planned_quantity)
    if default_value is None or default_value <= 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="planned_default (or planned_quantity) must be provided",
        )
    if quantity_mode == "RANGE":
        if planned_min is not None and planned_min > default_value:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="planned_min cannot exceed planned_default")
        if planned_max is not None and planned_max < default_value:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="planned_max cannot be lower than planned_default")


def _recipe_to_read(item: ServiceRecipeItem, product: LegacyProductCatalogItem | None) -> ServiceRecipeItemRead:
    effective_quantity = _coalesce_decimal(item.planned_default, item.planned_quantity) or Decimal("0")
    package_size_value = _coalesce_decimal(item.package_size_value, Decimal(str(product.quantity)) if product and product.quantity is not None else None)
    package_size_unit = (item.package_size_unit or "").strip().upper() or ("ML" if product and product.quantity is not None else None)
    package_unit_count = _coalesce_decimal(item.package_unit_count, Decimal("1") if package_size_value is not None else None)
    package_unit_label = (item.package_unit_label or "").strip().lower() or ("szt" if package_unit_count is not None else None)

    poj = None
    if package_size_value is not None:
        poj = f"{_format_decimal(package_size_value)} {(package_size_unit or '').lower()}".strip()
    iljedn = None
    if package_unit_count is not None:
        iljedn = f"{_format_decimal(package_unit_count)} {package_unit_label}".strip()
    total_label = None
    if effective_quantity > 0:
        total_label = f"{_format_decimal(effective_quantity)} {(item.recipe_unit_label or item.unit).lower()}"
        if poj:
            total_label = f"{total_label} × {poj}"
    return ServiceRecipeItemRead(
        id=item.id,
        service_id=item.service_id,
        variant_code=item.variant_code,
        position=item.position or 1,
        product_family=item.product_family,
        product_id=item.product_id,
        product_name=product.name if product else None,
        product_label_snapshot=item.product_label_snapshot or (product.name if product else None),
        is_optional=bool(getattr(item, "is_optional", False)),
        is_required=bool(getattr(item, "is_required", True)),
        quantity_mode=item.quantity_mode or "EXACT",
        planned_quantity=float(effective_quantity),
        planned_min=float(item.planned_min) if item.planned_min is not None else None,
        planned_default=float(item.planned_default) if item.planned_default is not None else None,
        planned_max=float(item.planned_max) if item.planned_max is not None else None,
        unit=item.recipe_unit_label or item.unit,
        recipe_unit_label=item.recipe_unit_label or item.unit,
        package_unit_count=float(package_unit_count) if package_unit_count is not None else None,
        package_unit_label=package_unit_label,
        package_size_value=float(package_size_value) if package_size_value is not None else None,
        package_size_unit=package_size_unit,
        inventory_mode=item.inventory_mode or "PER_SERVICE",
        note=item.note,
        poj=poj,
        iljedn=iljedn,
        total_label=total_label,
    )


def _normalize_excel_columns(columns: list[object]) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for raw in columns:
        source = str(raw or "").strip()
        if not source:
            continue
        key = source.lower()
        key = key.replace("ł", "l").replace("ó", "o").replace("ą", "a").replace("ś", "s").replace("ż", "z").replace("ź", "z").replace("ć", "c").replace("ń", "n").replace("ę", "e")
        key = " ".join(key.split())
        mapped[key] = source
    return mapped


def _as_optional_decimal(value: object) -> Decimal | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text or text.lower() == "nan":
        return None
    try:
        return Decimal(text)
    except Exception:
        return None


def _apply_excel_data_validations(writer: pd.ExcelWriter, row_count: int) -> None:
    workbook = writer.book
    worksheet = writer.sheets.get("recipes")
    if worksheet is None:
        return

    dictionaries_sheet = workbook.create_sheet("slowniki")
    dictionaries_sheet.sheet_state = "hidden"

    dictionaries_sheet["A1"] = "quantity_mode"
    for index, value in enumerate(QUANTITY_MODE_LABELS.values(), start=2):
        dictionaries_sheet[f"A{index}"] = value

    dictionaries_sheet["B1"] = "inventory_mode"
    for index, value in enumerate(INVENTORY_MODE_LABELS.values(), start=2):
        dictionaries_sheet[f"B{index}"] = value

    dictionaries_sheet["C1"] = "unit"
    for index, value in enumerate(UNIT_ALLOWED_VALUES, start=2):
        dictionaries_sheet[f"C{index}"] = value

    dictionaries_sheet["D1"] = "bool"
    for index, value in enumerate(BOOL_ALLOWED_VALUES, start=2):
        dictionaries_sheet[f"D{index}"] = value

    dictionaries_sheet["E1"] = "service_segment"
    for index, value in enumerate(SERVICE_SEGMENT_LABELS.values(), start=2):
        dictionaries_sheet[f"E{index}"] = value

    max_row = max(2, row_count + 1)
    header_to_column = {
        str(worksheet.cell(row=1, column=column_index).value or "").strip(): column_index
        for column_index in range(1, worksheet.max_column + 1)
    }

    def _attach_list_validation(header: str, formula: str, error_title: str, error_message: str) -> None:
        column_index = header_to_column.get(header)
        if column_index is None:
            return
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.errorTitle = error_title
        validation.error = error_message
        validation.promptTitle = "Dozwolone wartosci"
        validation.prompt = "Wybierz wartosc z listy."
        worksheet.add_data_validation(validation)
        validation.add(f"{column_letter}2:{column_letter}{max_row}")

    _attach_list_validation(
        "quantity_mode",
        "=slowniki!$A$2:$A$4",
        "Nieprawidlowy tryb ilosci",
        "Wybierz tryb z listy: Scisle|dokladnie, Zakres, Szacunkowo.",
    )
    _attach_list_validation(
        "inventory_mode",
        "=slowniki!$B$2:$B$4",
        "Nieprawidlowy sposob rozliczania",
        "Wybierz sposob rozliczania z listy.",
    )
    _attach_list_validation(
        "unit",
        "=slowniki!$C$2:$C$5",
        "Nieprawidlowa jednostka",
        "Wybierz jednostke z listy: PCS, G, ML, DOZA.",
    )
    _attach_list_validation(
        "recipe_unit_label",
        "=slowniki!$C$2:$C$5",
        "Nieprawidlowa jednostka operacyjna",
        "Wybierz jednostke operacyjna z listy: PCS, G, ML, DOZA.",
    )
    _attach_list_validation(
        "is_required",
        "=slowniki!$D$2:$D$3",
        "Nieprawidlowa wartosc logiczna",
        "Wybierz TRUE albo FALSE.",
    )
    _attach_list_validation(
        "is_optional",
        "=slowniki!$D$2:$D$3",
        "Nieprawidlowa wartosc logiczna",
        "Wybierz TRUE albo FALSE.",
    )
    _attach_list_validation(
        "service_segment",
        "=slowniki!$E$2:$E$5",
        "Nieprawidlowy znacznik uslugi",
        "Wybierz znacznik: Pani, Pan, Estetyka, Sprzedaz.",
    )


@router.get("/recipes/export/xlsx")
async def export_recipes_xlsx(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_manager_or_admin(current_user)
    rows = (
        db.query(ServiceRecipeItem, ServiceCatalogItem, LegacyProductCatalogItem)
        .join(ServiceCatalogItem, ServiceCatalogItem.id == ServiceRecipeItem.service_id)
        .outerjoin(LegacyProductCatalogItem, LegacyProductCatalogItem.id == ServiceRecipeItem.product_id)
        .order_by(ServiceCatalogItem.legacy_code.asc(), ServiceRecipeItem.position.asc(), ServiceRecipeItem.id.asc())
        .all()
    )
    data = []
    for recipe, service, product in rows:
        data.append(
            {
                "service_code": service.legacy_code,
                "service_name": service.name,
                "service_segment": SERVICE_SEGMENT_LABELS.get(_resolve_service_segment(service), _resolve_service_segment(service)),
                "position": recipe.position or 1,
                "product_family": recipe.product_family or "",
                "product_code": product.legacy_code if product else "",
                "product_name": recipe.product_label_snapshot or (product.name if product else ""),
                "is_required": bool(getattr(recipe, "is_required", True)),
                "is_optional": bool(getattr(recipe, "is_optional", False)),
                "quantity_mode": QUANTITY_MODE_LABELS.get(recipe.quantity_mode or "EXACT", recipe.quantity_mode or "EXACT"),
                "planned_quantity": float(recipe.planned_quantity) if recipe.planned_quantity is not None else "",
                "planned_min": float(recipe.planned_min) if recipe.planned_min is not None else "",
                "planned_default": float(recipe.planned_default) if recipe.planned_default is not None else "",
                "planned_max": float(recipe.planned_max) if recipe.planned_max is not None else "",
                "unit": recipe.unit or "PCS",
                "recipe_unit_label": recipe.recipe_unit_label or recipe.unit or "PCS",
                "package_unit_count": float(recipe.package_unit_count) if recipe.package_unit_count is not None else "",
                "package_unit_label": recipe.package_unit_label or "",
                "package_size_value": float(recipe.package_size_value) if recipe.package_size_value is not None else "",
                "package_size_unit": recipe.package_size_unit or "",
                "inventory_mode": INVENTORY_MODE_LABELS.get(recipe.inventory_mode or "PER_SERVICE", recipe.inventory_mode or "PER_SERVICE"),
                "note": recipe.note or "",
            }
        )

    df = pd.DataFrame(data)
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="recipes")
        _apply_excel_data_validations(writer, len(df.index))
    out.seek(0)
    filename = f"service_recipes_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/recipes/import/xlsx", response_model=RecipeImportResponse)
async def import_recipes_xlsx(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_manager_or_admin(current_user)
    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .xlsx or .xls files are supported")

    payload = await file.read()
    if not payload:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Uploaded file is empty")

    try:
        df = pd.read_excel(BytesIO(payload), sheet_name=0)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Cannot parse Excel file: {exc}") from exc
    if df.empty:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Excel has no rows")

    colmap = _normalize_excel_columns(list(df.columns))
    required_aliases = {
        "service_code": ["service_code", "kod uslugi", "kod usługi"],
        "position": ["position", "pozycja", "poz"],
        "product_family": ["product_family", "rodzina", "rodzina produktu"],
        "planned_quantity": ["planned_quantity", "planned_default", "domyslna", "ilość jednostek do receptury", "ilosc jednostek do receptury"],
    }
    resolved: dict[str, str] = {}
    for target, aliases in required_aliases.items():
        match = next((colmap.get(alias) for alias in aliases if alias in colmap), None)
        if match is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Missing required column: {target}")
        resolved[target] = match

    service_code_col = resolved["service_code"]
    position_col = resolved["position"]
    product_family_col = resolved["product_family"]
    planned_quantity_col = resolved["planned_quantity"]

    product_code_col = next((colmap.get(alias) for alias in ["product_code", "kod produktu"] if alias in colmap), None)
    product_name_col = next((colmap.get(alias) for alias in ["product_name", "nazwa produktu"] if alias in colmap), None)
    quantity_mode_col = next((colmap.get(alias) for alias in ["quantity_mode", "tryb ilosci"] if alias in colmap), None)
    unit_col = next((colmap.get(alias) for alias in ["unit", "jm", "jednostka"] if alias in colmap), None)
    recipe_unit_label_col = next((colmap.get(alias) for alias in ["recipe_unit_label", "jednostka operacyjna"] if alias in colmap), None)
    package_unit_count_col = next((colmap.get(alias) for alias in ["package_unit_count", "ilosc klientow z opakowania"] if alias in colmap), None)
    package_unit_label_col = next((colmap.get(alias) for alias in ["package_unit_label", "iljedn"] if alias in colmap), None)
    package_size_value_col = next((colmap.get(alias) for alias in ["package_size_value"] if alias in colmap), None)
    package_size_unit_col = next((colmap.get(alias) for alias in ["package_size_unit"] if alias in colmap), None)
    inventory_mode_col = next((colmap.get(alias) for alias in ["inventory_mode", "sposob rozliczania"] if alias in colmap), None)
    service_segment_col = next((colmap.get(alias) for alias in ["service_segment", "znacznik", "segment uslugi", "segment usługi"] if alias in colmap), None)
    note_col = next((colmap.get(alias) for alias in ["note", "notatka"] if alias in colmap), None)

    service_by_code = {
        row.legacy_code: row
        for row in db.query(ServiceCatalogItem).all()
    }
    service_by_id = {
        row.id: row
        for row in service_by_code.values()
    }
    product_by_code = {
        row.legacy_code.strip().lower(): row
        for row in db.query(LegacyProductCatalogItem).all()
        if row.legacy_code
    }
    product_by_name = {
        (row.name or "").strip().lower(): row
        for row in db.query(LegacyProductCatalogItem).all()
        if row.name
    }

    grouped: dict[int, list[ServiceRecipeItem]] = {}
    service_segment_updates: dict[int, str] = {}
    imported_rows = 0
    skipped_rows = 0
    unresolved_products = 0

    for _, row in df.iterrows():
        service_code = str(row.get(service_code_col, "") or "").strip()
        if not service_code:
            skipped_rows += 1
            continue
        if service_code.isdigit():
            service_code = f"{int(service_code):04d}"
        service = service_by_code.get(service_code)
        if service is None:
            skipped_rows += 1
            continue
        if service_segment_col:
            raw_segment = str(row.get(service_segment_col, "") or "").strip()
            if raw_segment:
                normalized_segment = _normalize_service_segment(raw_segment)
                existing = service_segment_updates.get(service.id)
                if existing and existing != normalized_segment:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Conflicting service_segment values for service_code={service_code}",
                    )
                service_segment_updates[service.id] = normalized_segment

        position_raw = str(row.get(position_col, "") or "").strip()
        try:
            position = int(float(position_raw)) if position_raw else 1
        except Exception:
            position = 1
        product_family = str(row.get(product_family_col, "") or "").strip()
        if not product_family:
            skipped_rows += 1
            continue
        planned_quantity = _as_optional_decimal(row.get(planned_quantity_col))
        if planned_quantity is None or planned_quantity <= 0:
            skipped_rows += 1
            continue

        product = None
        if product_code_col:
            code = str(row.get(product_code_col, "") or "").strip().lower()
            if code:
                product = product_by_code.get(code)
        if product is None and product_name_col:
            pname = str(row.get(product_name_col, "") or "").strip().lower()
            if pname:
                product = product_by_name.get(pname)
                if product is None:
                    unresolved_products += 1

        quantity_mode = _normalize_quantity_mode(str(row.get(quantity_mode_col, "") or "EXACT")) if quantity_mode_col else "EXACT"
        inventory_mode = _normalize_inventory_mode(str(row.get(inventory_mode_col, "") or "PER_SERVICE")) if inventory_mode_col else "PER_SERVICE"
        unit = str(row.get(unit_col, "") or "PCS").strip().upper() if unit_col else "PCS"
        recipe_unit_label = (
            str(row.get(recipe_unit_label_col, "") or "").strip().upper()
            if recipe_unit_label_col
            else unit
        ) or unit
        package_unit_count = _as_optional_decimal(row.get(package_unit_count_col)) if package_unit_count_col else None
        package_unit_label = str(row.get(package_unit_label_col, "") or "").strip().lower() if package_unit_label_col else None
        package_size_value = _as_optional_decimal(row.get(package_size_value_col)) if package_size_value_col else None
        package_size_unit = str(row.get(package_size_unit_col, "") or "").strip().upper() if package_size_unit_col else None
        note = str(row.get(note_col, "") or "").strip() if note_col else ""

        recipe_row = ServiceRecipeItem(
            service_id=service.id,
            position=max(1, position),
            product_family=product_family,
            product_id=product.id if product else None,
            product_label_snapshot=(str(row.get(product_name_col, "") or "").strip() if product_name_col else "") or (product.name if product else None),
            is_optional=False,
            is_required=True,
            quantity_mode=quantity_mode,
            planned_quantity=planned_quantity,
            planned_default=planned_quantity,
            planned_min=None,
            planned_max=None,
            unit=unit,
            recipe_unit_label=recipe_unit_label,
            package_unit_count=package_unit_count,
            package_unit_label=package_unit_label or None,
            package_size_value=package_size_value,
            package_size_unit=package_size_unit or None,
            inventory_mode=inventory_mode,
            note=note or None,
        )
        grouped.setdefault(service.id, []).append(recipe_row)
        imported_rows += 1

    for service_id, recipe_rows in grouped.items():
        db.query(ServiceRecipeItem).filter(ServiceRecipeItem.service_id == service_id).delete(synchronize_session=False)
        recipe_rows.sort(key=lambda item: (item.position or 1, item.product_label_snapshot or ""))
        for idx, recipe_row in enumerate(recipe_rows, start=1):
            recipe_row.position = idx
            db.add(recipe_row)
    for service_id, normalized_segment in service_segment_updates.items():
        service = service_by_id.get(service_id)
        if service is None:
            continue
        service.service_type_code = normalized_segment
    db.commit()

    return RecipeImportResponse(
        imported_services=len(grouped),
        imported_rows=imported_rows,
        skipped_rows=skipped_rows,
        unresolved_products=unresolved_products,
    )


@router.get("/services/{service_id}/recipe", response_model=list[ServiceRecipeItemRead])
async def list_service_recipe(
    service_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    del current_user
    if not db.query(ServiceCatalogItem.id).filter(ServiceCatalogItem.id == service_id).first():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Service not found")
    items = (
        db.query(ServiceRecipeItem)
        .filter(ServiceRecipeItem.service_id == service_id)
        .order_by(ServiceRecipeItem.position.asc(), ServiceRecipeItem.id.asc())
        .all()
    )
    product_ids = [item.product_id for item in items if item.product_id is not None]
    products_by_id = {}
    if product_ids:
        products_by_id = {
            row.id: row
            for row in db.query(LegacyProductCatalogItem).filter(LegacyProductCatalogItem.id.in_(product_ids)).all()
        }
    return [_recipe_to_read(item, products_by_id.get(item.product_id)) for item in items]


@router.post("/services/{service_id}/recipe", response_model=ServiceRecipeItemRead, status_code=status.HTTP_201_CREATED)
async def create_service_recipe_item(
    service_id: int,
    payload: ServiceRecipeItemCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_manager_or_admin(current_user)
    if not db.query(ServiceCatalogItem.id).filter(ServiceCatalogItem.id == service_id).first():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Service not found")
    product = None
    if payload.product_id is not None:
        product = db.query(LegacyProductCatalogItem).filter(LegacyProductCatalogItem.id == payload.product_id).first()
        if not product:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
    quantity_mode = _normalize_quantity_mode(payload.quantity_mode)
    inventory_mode = _normalize_inventory_mode(payload.inventory_mode)
    planned_quantity = _coerce_decimal(payload.planned_quantity)
    planned_min = _coerce_decimal(payload.planned_min)
    planned_default = _coerce_decimal(payload.planned_default) or planned_quantity
    planned_max = _coerce_decimal(payload.planned_max)
    product_family = (payload.product_family or "").strip() or None
    _validate_recipe_payload(product_family, quantity_mode, planned_quantity, planned_min, planned_default, planned_max)
    item = ServiceRecipeItem(
        service_id=service_id,
        variant_code=(payload.variant_code or "").strip().upper() or None,
        position=payload.position or 1,
        product_family=product_family,
        product_id=payload.product_id,
        product_label_snapshot=(payload.product_label_snapshot or "").strip() or (product.name if product else None),
        is_optional=payload.is_optional,
        is_required=payload.is_required,
        quantity_mode=quantity_mode,
        planned_quantity=planned_quantity or planned_default or Decimal("0"),
        planned_min=planned_min,
        planned_default=planned_default,
        planned_max=planned_max,
        unit=payload.unit.strip().upper(),
        recipe_unit_label=(payload.recipe_unit_label or payload.unit).strip().upper(),
        package_unit_count=_coerce_decimal(payload.package_unit_count),
        package_unit_label=(payload.package_unit_label or "").strip().lower() or None,
        package_size_value=_coerce_decimal(payload.package_size_value),
        package_size_unit=(payload.package_size_unit or "").strip().upper() or None,
        inventory_mode=inventory_mode,
        note=(payload.note or "").strip() or None,
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return _recipe_to_read(item, product)


@router.patch("/services/{service_id}/recipe/{recipe_id}", response_model=ServiceRecipeItemRead)
async def update_service_recipe_item(
    service_id: int,
    recipe_id: int,
    payload: ServiceRecipeItemUpdate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_manager_or_admin(current_user)
    item = (
        db.query(ServiceRecipeItem)
        .filter(ServiceRecipeItem.id == recipe_id, ServiceRecipeItem.service_id == service_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Recipe item not found")
    product = None
    provided = payload.model_fields_set
    if "product_id" in provided:
        if payload.product_id is not None:
            product = db.query(LegacyProductCatalogItem).filter(LegacyProductCatalogItem.id == payload.product_id).first()
            if not product:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Product not found")
        item.product_id = payload.product_id
    elif item.product_id is not None:
        product = db.query(LegacyProductCatalogItem).filter(LegacyProductCatalogItem.id == item.product_id).first()
    if "variant_code" in provided:
        item.variant_code = (payload.variant_code or "").strip().upper() or None
    if "position" in provided and payload.position is not None:
        item.position = payload.position
    if "product_family" in provided:
        item.product_family = (payload.product_family or "").strip() or None
    if "product_label_snapshot" in provided:
        item.product_label_snapshot = (payload.product_label_snapshot or "").strip() or (product.name if product else None)
    if "planned_quantity" in provided and payload.planned_quantity is not None:
        item.planned_quantity = Decimal(str(payload.planned_quantity))
        if item.planned_default is None:
            item.planned_default = item.planned_quantity
    if "planned_min" in provided:
        item.planned_min = _coerce_decimal(payload.planned_min)
    if "planned_default" in provided:
        item.planned_default = _coerce_decimal(payload.planned_default)
    if "planned_max" in provided:
        item.planned_max = _coerce_decimal(payload.planned_max)
    if "is_optional" in provided and payload.is_optional is not None:
        item.is_optional = payload.is_optional
    if "is_required" in provided and payload.is_required is not None:
        item.is_required = payload.is_required
    if "quantity_mode" in provided and payload.quantity_mode is not None:
        item.quantity_mode = _normalize_quantity_mode(payload.quantity_mode)
    if "unit" in provided and payload.unit is not None:
        item.unit = payload.unit.strip().upper()
    if "recipe_unit_label" in provided:
        item.recipe_unit_label = (payload.recipe_unit_label or "").strip().upper() or None
    if "package_unit_count" in provided:
        item.package_unit_count = _coerce_decimal(payload.package_unit_count)
    if "package_unit_label" in provided:
        item.package_unit_label = (payload.package_unit_label or "").strip().lower() or None
    if "package_size_value" in provided:
        item.package_size_value = _coerce_decimal(payload.package_size_value)
    if "package_size_unit" in provided:
        item.package_size_unit = (payload.package_size_unit or "").strip().upper() or None
    if "inventory_mode" in provided and payload.inventory_mode is not None:
        item.inventory_mode = _normalize_inventory_mode(payload.inventory_mode)
    if "note" in provided:
        item.note = (payload.note or "").strip() or None
    _validate_recipe_payload(
        item.product_family,
        item.quantity_mode or "EXACT",
        item.planned_quantity,
        item.planned_min,
        item.planned_default or item.planned_quantity,
        item.planned_max,
    )
    db.commit()
    db.refresh(item)
    return _recipe_to_read(item, product)


@router.delete("/services/{service_id}/recipe/{recipe_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_service_recipe_item(
    service_id: int,
    recipe_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    _ensure_manager_or_admin(current_user)
    item = (
        db.query(ServiceRecipeItem)
        .filter(ServiceRecipeItem.id == recipe_id, ServiceRecipeItem.service_id == service_id)
        .first()
    )
    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Recipe item not found")
    db.delete(item)
    db.commit()
    return None


@router.post("/appointments/{appointment_id}/generate-issues", response_model=GeneratedIssueResponse)
async def generate_inventory_issues_for_appointment(
    appointment_id: int,
    current_user: User = Depends(get_current_user),
    current_staff: StaffMember | None = Depends(get_current_staff_member),
    db: Session = Depends(get_db),
):
    appointment = db.query(Appointment).filter(Appointment.id == appointment_id).first()
    if not appointment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Appointment not found")
    if (appointment.status or "").lower() != "done":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Issues can be generated only for done appointments")

    _ensure_issue_generation_access(db, current_user, current_staff, appointment)

    stock_location = (
        db.query(StockLocation)
        .filter(
            StockLocation.salon_id == appointment.salon_id,
            StockLocation.is_active.is_(True),
            StockLocation.location_type.in_(["BACKBAR", "MIXED"]),
        )
        .order_by(StockLocation.location_type.asc(), StockLocation.code.asc())
        .first()
    )
    if not stock_location:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="No active BACKBAR/MIXED stock location in salon")

    performed_lines = (
        db.query(AppointmentPerformedLine)
        .filter(AppointmentPerformedLine.appointment_id == appointment_id)
        .order_by(AppointmentPerformedLine.id.asc())
        .all()
    )
    if not performed_lines:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Appointment has no performed lines")

    service_ids = sorted({line.service_id for line in performed_lines})
    recipes_by_service: dict[int, list[ServiceRecipeItem]] = {}
    if service_ids:
        for item in (
            db.query(ServiceRecipeItem)
            .filter(ServiceRecipeItem.service_id.in_(service_ids))
            .order_by(ServiceRecipeItem.id.asc())
            .all()
        ):
            recipes_by_service.setdefault(item.service_id, []).append(item)

    product_ids = {item.product_id for rows in recipes_by_service.values() for item in rows if item.product_id is not None}
    products_by_id = {}
    if product_ids:
        products_by_id = {
            row.id: row
            for row in db.query(LegacyProductCatalogItem).filter(LegacyProductCatalogItem.id.in_(product_ids)).all()
        }

    issue_ids: list[int] = []
    for performed_line in performed_lines:
        recipe_items = recipes_by_service.get(performed_line.service_id, [])
        if not recipe_items:
            continue

        existing_issue = (
            db.query(InventoryIssue)
            .filter(InventoryIssue.performed_line_id == performed_line.id)
            .order_by(InventoryIssue.id.asc())
            .first()
        )
        if existing_issue:
            issue_ids.append(existing_issue.id)
            continue

        issue = InventoryIssue(
            salon_id=appointment.salon_id,
            stock_location_id=stock_location.id,
            appointment_id=appointment.id,
            service_id=performed_line.service_id,
            performed_line_id=performed_line.id,
            staff_id=performed_line.worker_id,
            issue_time=performed_line.performed_at,
            status="PLANNED",
            remarks="Wygenerowano z receptury uslugi",
        )
        db.add(issue)
        db.flush()

        for recipe_item in recipe_items:
            product = products_by_id.get(recipe_item.product_id) if recipe_item.product_id is not None else None
            db.add(
                InventoryIssueLine(
                    inventory_issue_id=issue.id,
                    appointment_id=appointment.id,
                    service_id=performed_line.service_id,
                    recipe_item_id=recipe_item.id,
                    performed_line_id=performed_line.id,
                    product_id=recipe_item.product_id,
                    quantity_planned=recipe_item.planned_quantity,
                    quantity_actual=None,
                    unit=recipe_item.unit,
                    unit_cost=_estimate_unit_cost(product),
                    total_cost=Decimal("0"),
                )
            )

        issue_ids.append(issue.id)

    db.commit()
    return GeneratedIssueResponse(issue_ids=issue_ids)
